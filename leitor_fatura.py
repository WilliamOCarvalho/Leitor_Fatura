import argparse
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Tuple

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# -----------------------------
# Config / Keywords
# -----------------------------

DEFAULT_KEYWORDS_FILE = "keywords.json"


def load_keywords(path: Path) -> List[str]:
    if not path.exists():
        return ["UBER", "99"]
    data = json.loads(path.read_text(encoding="utf-8"))
    kws = data.get("keywords", [])
    return [str(k).strip() for k in kws if str(k).strip()]


def save_keywords(path: Path, keywords: List[str]) -> None:
    path.write_text(
        json.dumps({"keywords": keywords}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def add_keyword(path: Path, keyword: str) -> None:
    keyword = keyword.strip()
    if not keyword:
        raise ValueError("Keyword vazia.")
    kws = load_keywords(path)
    # padroniza em caixa alta pra bater melhor
    up = keyword.upper()
    if up not in [k.upper() for k in kws]:
        kws.append(keyword)
        save_keywords(path, kws)


def remove_keyword(path: Path, keyword: str) -> None:
    keyword = keyword.strip()
    if not keyword:
        raise ValueError("Keyword vazia.")
    kws = load_keywords(path)
    target = keyword.upper()
    filtered = [k for k in kws if k.upper() != target]
    save_keywords(path, filtered)


# -----------------------------
# Parsing helpers
# -----------------------------

# Datas comuns em fatura BR: 31/01 ou 31/01/2026
DATE_RE = re.compile(r"\b(\d{2}/\d{2})(?:/(\d{2,4}))?\b")

# Valores: 12,34  |  1.234,56  |  -12,34  |  (12,34)
AMOUNT_RE = re.compile(
    r"(?P<neg_par>\()?"
    r"(?P<sign>-)?"
    r"(?P<num>\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2})"
    r"(?P<neg_par2>\))?"
)

CURRENCY_HINTS = ["R$", "BRL"]


def br_money_to_float(s: str) -> float:
    s = s.strip()
    is_paren = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    s = s.replace("R$", "").replace("BRL", "").strip()
    s = s.replace(".", "").replace(",", ".")
    val = float(s)
    if is_paren:
        val *= -1
    return val


def normalize_text(s: str) -> str:
    # facilita match (sem mexer em acentos, só normaliza espaços)
    return re.sub(r"[ \t]+", " ", s).strip()


@dataclass
class Lancamento:
    app: str
    data: str  # "dd/mm" ou "dd/mm/aaaa" (como veio)
    descricao: str
    valor: float
    pagina: int


def find_best_date(line: str) -> Optional[str]:
    m = DATE_RE.search(line)
    if not m:
        return None
    ddmm = m.group(1)
    yyyy = m.group(2)
    if yyyy:
        if len(yyyy) == 2:
            # heurística: 20xx
            yyyy = "20" + yyyy
        return f"{ddmm}/{yyyy}"
    return ddmm


def find_best_amount(line: str) -> Optional[float]:
    # pega o último valor da linha (normalmente o valor do lançamento fica no fim)
    matches = list(AMOUNT_RE.finditer(line))
    if not matches:
        return None
    m = matches[-1]
    raw = m.group(0)
    # evita capturar coisas que sejam tipo "10/12" (não casa com AMOUNT_RE, ok)
    try:
        return br_money_to_float(raw)
    except Exception:
        return None


def guess_app_from_line(line: str, keywords: List[str]) -> Optional[str]:
    up = line.upper()
    for kw in keywords:
        if kw.upper() in up:
            # “nome” que aparece na planilha (pode ser a própria keyword)
            # Se quiser padronizar Uber/99, dá pra ajustar aqui.
            return kw.upper()
    return None


def extract_candidates_from_text(text: str, page_index: int, keywords: List[str]) -> List[Lancamento]:
    results: List[Lancamento] = []
    for raw_line in text.splitlines():
        line = normalize_text(raw_line)
        if not line:
            continue

        app = guess_app_from_line(line, keywords)
        if not app:
            continue

        data = find_best_date(line)
        valor = find_best_amount(line)

        # Heurística: se não veio data/valor na mesma linha, tenta “colar” o que dá.
        # Aqui mantemos simples e só registra se achar valor.
        if valor is None:
            continue

        # descrição = linha sem o valor final (tentativa)
        descricao = line
        # remove o último valor detectado do fim, se estiver no fim
        # (ajuda a deixar a descrição mais limpa)
        last_amount = AMOUNT_RE.findall(line)
        if last_amount:
            # remove apenas a última ocorrência numérica
            descricao = AMOUNT_RE.sub("", descricao, count=1)  # remove a primeira ocorrência
            # como removemos a primeira, nem sempre é a última.
            # então fazemos um “jeitinho”: se ficou ruim, mantém original.
            if len(descricao.strip()) < 3:
                descricao = line

        results.append(
            Lancamento(
                app=app,
                data=data or "",
                descricao=line,
                valor=valor,
                pagina=page_index + 1,
            )
        )
    return results


def read_pdf_extract(pdf_path: Path, keywords: List[str]) -> List[Lancamento]:
    lancs: List[Lancamento] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            if not text.strip():
                continue
            lancs.extend(extract_candidates_from_text(text, i, keywords))
    return lancs


# -----------------------------
# Excel output
# -----------------------------

def autosize_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def write_xlsx(output_path: Path, lancs: List[Lancamento], keywords: List[str]) -> None:
    wb = Workbook()

    # Aba Lancamentos
    ws = wb.active
    ws.title = "Lancamentos"

    headers = ["App/Termo", "Data", "Descrição", "Valor (R$)", "Página"]
    ws.append(headers)

    header_fill = PatternFill(start_color="FFEEEEEE", end_color="FFEEEEEE", fill_type="solid")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for l in lancs:
        ws.append([l.app, l.data, l.descricao, l.valor, l.pagina])

    # formatação básica
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=4).number_format = '"R$" #,##0.00;[Red]\-"R$" #,##0.00'

    autosize_columns(ws)

    # Aba Resumo
    ws2 = wb.create_sheet("Resumo")
    ws2.append(["Termo", "Total (R$)"])
    for c in range(1, 3):
        cell = ws2.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    totals: Dict[str, float] = {}
    for l in lancs:
        totals[l.app] = totals.get(l.app, 0.0) + l.valor

    # garante que apareçam todos os termos do keywords, mesmo que 0
    for kw in keywords:
        key = kw.upper()
        ws2.append([key, round(totals.get(key, 0.0), 2)])

    total_geral = sum(totals.values())
    ws2.append(["TOTAL GERAL", round(total_geral, 2)])

    # formata valores
    for row in range(2, ws2.max_row + 1):
        ws2.cell(row=row, column=2).number_format = '"R$" #,##0.00;[Red]\-"R$" #,##0.00'
        if ws2.cell(row=row, column=1).value == "TOTAL GERAL":
            ws2.cell(row=row, column=1).font = Font(bold=True)
            ws2.cell(row=row, column=2).font = Font(bold=True)

    autosize_columns(ws2)

    wb.save(str(output_path))


# -----------------------------
# CLI
# -----------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Leitor de fatura PDF: encontra gastos (Uber/99 etc.) e gera planilha com totais."
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_run = sub.add_parser("run", help="Ler PDF e gerar XLSX")
    p_run.add_argument("pdf", type=str, help="Caminho do PDF da fatura")
    p_run.add_argument("-k", "--keywords", type=str, default=DEFAULT_KEYWORDS_FILE, help="Arquivo keywords.json")
    p_run.add_argument("-o", "--output", type=str, default="resultado_fatura.xlsx", help="Saída .xlsx")

    p_list = sub.add_parser("list", help="Listar termos configurados")
    p_list.add_argument("-k", "--keywords", type=str, default=DEFAULT_KEYWORDS_FILE)

    p_add = sub.add_parser("add", help="Adicionar termo para procurar")
    p_add.add_argument("term", type=str, help="Termo (ex: UBER*TRIP ou 99APP)")
    p_add.add_argument("-k", "--keywords", type=str, default=DEFAULT_KEYWORDS_FILE)

    p_rm = sub.add_parser("remove", help="Remover termo")
    p_rm.add_argument("term", type=str)
    p_rm.add_argument("-k", "--keywords", type=str, default=DEFAULT_KEYWORDS_FILE)

    args = parser.parse_args()
    kw_path = Path(getattr(args, "keywords", DEFAULT_KEYWORDS_FILE))

    if args.cmd == "list":
        kws = load_keywords(kw_path)
        print("Termos configurados:")
        for k in kws:
            print(f" - {k}")
        return

    if args.cmd == "add":
        add_keyword(kw_path, args.term)
        print(f"Adicionado: {args.term}")
        return

    if args.cmd == "remove":
        remove_keyword(kw_path, args.term)
        print(f"Removido: {args.term}")
        return

    if args.cmd == "run":
        pdf_path = Path(args.pdf)
        if not pdf_path.exists():
            raise FileNotFoundError(f"PDF não encontrado: {pdf_path}")

        keywords = load_keywords(kw_path)
        lancs = read_pdf_extract(pdf_path, keywords)

        out_path = Path(args.output)
        write_xlsx(out_path, lancs, keywords)

        total = sum(l.valor for l in lancs)
        print(f"Encontrados {len(lancs)} lançamentos. Total geral: R$ {total:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        print(f"Planilha gerada em: {out_path.resolve()}")
        return


if __name__ == "__main__":
    main()
